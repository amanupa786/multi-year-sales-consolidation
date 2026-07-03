# ============================================================
# pipeline.py
# Multi-Year Sales Consolidation (2013–2024)
# ============================================================
# PURPOSE : One-command pipeline that:
#             1. Loads all raw Excel files
#             2. Cleans and standardises each file
#             3. Consolidates into a single CSV
#             4. Produces YoY trend analysis Excel
#             5. Produces a multi-sheet summary report
#             6. Saves professional charts (PNG)
#             7. Loads data into MySQL
# RUN     : python pipeline.py
# PRE-REQ : Run generate_data.py first to create raw files.
# ============================================================

import os
import sys
import glob
import warnings
import traceback
import numpy as np
import pandas as pd
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
matplotlib.use("Agg")   # non-interactive backend for saving PNGs

# ── Import project config ────────────────────────────────────
try:
    from config import DB_CONFIG, RAW_DATA_PATH, PROCESSED_DATA_PATH, OUTPUTS_PATH, CHARTS_PATH
except ImportError:
    print("[ERROR] config.py not found. Please ensure config.py is in the project root.")
    sys.exit(1)

# ── Ensure output directories exist ─────────────────────────
os.makedirs(PROCESSED_DATA_PATH, exist_ok=True)
os.makedirs(OUTPUTS_PATH,        exist_ok=True)
os.makedirs(CHARTS_PATH,         exist_ok=True)


# ============================================================
# UTILITY HELPERS
# ============================================================

def section_banner(title: str):
    """Print a formatted section banner to the terminal."""
    print("\n" + "=" * 60)
    print(f"  {title}")
    print("=" * 60)


def apply_excel_formatting(ws, header_fill_hex: str = "1F4E79"):
    """
    Apply consistent Excel formatting to a worksheet:
      - Bold white header row with colored background
      - Auto column width
      - Center-align header cells
    """
    header_fill  = PatternFill(fill_type="solid", fgColor=header_fill_hex)
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    border_side  = Side(style="thin", color="CCCCCC")
    cell_border  = Border(
        left=border_side, right=border_side,
        top=border_side,  bottom=border_side
    )

    for col_idx, cell in enumerate(ws[1], start=1):
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"


def fmt_currency(ws, col_letters: list, start_row: int = 2):
    """Apply accounting number format to given columns."""
    for col_letter in col_letters:
        for row in ws.iter_rows(
            min_row=start_row, max_row=ws.max_row,
            min_col=ws[col_letter + "1"].column,
            max_col=ws[col_letter + "1"].column
        ):
            for cell in row:
                cell.number_format = '#,##0.00'


# ============================================================
# STEP 1 — LOAD ALL RAW FILES
# ============================================================

def step1_load_files(raw_path: str) -> list[tuple[int, pd.DataFrame]]:
    """
    Detect all .xlsx files in raw_path, load each into a DataFrame,
    and tag with a Year column extracted from the filename.
    Returns a list of (year, dataframe) tuples.
    """
    section_banner("STEP 1 — Loading Raw Excel Files")

    pattern = os.path.join(raw_path, "sales_*.xlsx")
    files   = sorted(glob.glob(pattern))

    if not files:
        print(f"[ERROR] No .xlsx files found in '{raw_path}'.")
        print("        Please run generate_data.py first.")
        sys.exit(1)

    yearly_dfs = []
    for filepath in files:
        filename = os.path.basename(filepath)
        try:
            year = int(filename.replace("sales_", "").replace(".xlsx", ""))
        except ValueError:
            print(f"  [SKIP] Cannot parse year from filename: {filename}")
            continue

        df = pd.read_excel(filepath, engine="openpyxl")
        df["Year"] = year
        yearly_dfs.append((year, df))
        print(f"  [OK]  {filename}  →  {len(df):,} rows loaded")

    print(f"\n  Total files loaded: {len(yearly_dfs)}")
    return yearly_dfs


# ============================================================
# STEP 2 — DATA CLEANING
# ============================================================

def step2_clean(year: int, df: pd.DataFrame, quality_log: list) -> pd.DataFrame:
    """
    Clean a single year's DataFrame. Returns the cleaned DataFrame.
    All issues found are appended to quality_log for the Data Quality sheet.
    """
    issues = {
        "Year":              year,
        "Raw_Rows":          len(df),
        "Missing_Discount":  0,
        "Missing_ShipCost":  0,
        "Missing_Region":    0,
        "Duplicates_Removed":0,
        "Zero_Null_Sales":   0,
        "Sales_Mismatches":  0,
        "Clean_Rows":        0,
    }

    # 2a. Standardise column names: strip whitespace, title case
    df.columns = [c.strip().title().replace(" ", "_") for c in df.columns]

    # Ensure expected columns exist
    expected_cols = [
        "Order_Id", "Order_Date", "Customer_Name", "Customer_Segment",
        "Region", "Country", "Category", "Sub_Category", "Product_Name",
        "Quantity", "Unit_Price", "Discount", "Sales", "Profit",
        "Shipping_Cost", "Ship_Mode", "Year"
    ]
    # Rename Order_ID variants
    df.rename(columns={"Order_Id": "Order_ID"}, inplace=True)

    # 2b. Fix inconsistent text casing
    for col in ["Category", "Sub_Category", "Region", "Customer_Segment", "Ship_Mode", "Country"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip().str.title()
            df[col] = df[col].replace("Nan", np.nan)

    # 2c. Convert Order_Date to proper datetime
    df["Order_Date"] = pd.to_datetime(df["Order_Date"], errors="coerce")

    # 2d. Fill missing Discount with 0
    issues["Missing_Discount"] = int(df["Discount"].isna().sum())
    df["Discount"] = df["Discount"].fillna(0)

    # 2e. Fill missing Shipping_Cost with median for that year
    issues["Missing_ShipCost"] = int(df["Shipping_Cost"].isna().sum())
    shipping_median = df["Shipping_Cost"].median()
    df["Shipping_Cost"] = df["Shipping_Cost"].fillna(shipping_median)

    # 2f. Fill missing Region with mode (most common)
    issues["Missing_Region"] = int(df["Region"].isna().sum())
    if df["Region"].notna().any():
        region_mode = df["Region"].mode()[0]
        df["Region"] = df["Region"].fillna(region_mode)

    # 2g. Remove duplicate rows
    before = len(df)
    df.drop_duplicates(inplace=True)
    issues["Duplicates_Removed"] = before - len(df)

    # 2h. Remove rows where Sales is null or zero
    before = len(df)
    df = df[df["Sales"].notna() & (df["Sales"] != 0)]
    issues["Zero_Null_Sales"] = before - len(df)

    # 2i. Validate Sales = Quantity × Unit_Price × (1 – Discount)
    #     Flag mismatches (>1% tolerance) but do NOT remove — just note them
    df["Sales_Calculated"] = (
        df["Quantity"] * df["Unit_Price"] * (1 - df["Discount"])
    ).round(2)
    mismatch_mask = (
        (df["Sales_Calculated"] - df["Sales"]).abs() / df["Sales"].abs() > 0.01
    )
    issues["Sales_Mismatches"] = int(mismatch_mask.sum())
    df.drop(columns=["Sales_Calculated"], inplace=True)

    # 2j. Add derived columns
    df["Profit_Margin_%"] = np.where(
        df["Sales"] != 0,
        (df["Profit"] / df["Sales"] * 100).round(2),
        0
    )
    df["Revenue_Band"] = pd.cut(
        df["Sales"],
        bins=[-np.inf, 500, 2000, np.inf],
        labels=["Low", "Medium", "High"]
    )

    issues["Clean_Rows"] = len(df)
    quality_log.append(issues)

    return df


def step2_clean_all(yearly_dfs: list) -> tuple[list, list]:
    """Run cleaning on every yearly DataFrame. Returns cleaned list + quality log."""
    section_banner("STEP 2 — Cleaning & Standardising Data")

    quality_log   = []
    cleaned_dfs   = []

    for year, df in yearly_dfs:
        print(f"  Cleaning {year}...", end=" ", flush=True)
        cleaned = step2_clean(year, df, quality_log)
        cleaned_dfs.append(cleaned)
        log = quality_log[-1]
        print(
            f"Done. Rows: {log['Raw_Rows']} → {log['Clean_Rows']}  "
            f"(dupes: {log['Duplicates_Removed']}, "
            f"missing discount: {log['Missing_Discount']}, "
            f"missing region: {log['Missing_Region']})"
        )

    return cleaned_dfs, quality_log


# ============================================================
# STEP 3 — CONSOLIDATE
# ============================================================

def step3_consolidate(cleaned_dfs: list) -> pd.DataFrame:
    """Merge all cleaned DataFrames and save to consolidated_sales.csv."""
    section_banner("STEP 3 — Consolidating Into Master Dataset")

    master_df = pd.concat(cleaned_dfs, ignore_index=True)

    # Sort by date
    master_df.sort_values(["Year", "Order_Date"], inplace=True)
    master_df.reset_index(drop=True, inplace=True)

    output_path = os.path.join(PROCESSED_DATA_PATH, "consolidated_sales.csv")
    master_df.to_csv(output_path, index=False)

    print(f"  Total rows    : {len(master_df):,}")
    print(f"  Total columns : {len(master_df.columns)}")
    print(f"  Date range    : {master_df['Order_Date'].min().date()}  →  {master_df['Order_Date'].max().date()}")
    print(f"  Years covered : {sorted(master_df['Year'].unique())}")
    print(f"  Saved to      : {output_path}")

    return master_df


# ============================================================
# STEP 4 — YEAR-OVER-YEAR ANALYSIS
# ============================================================

def step4_yoy_analysis(master_df: pd.DataFrame) -> pd.DataFrame:
    """
    Compute per-year KPIs and YoY growth rates.
    Saves result to outputs/yoy_trend_analysis.xlsx with formatting.
    """
    section_banner("STEP 4 — Year-over-Year Trend Analysis")

    # Per-year aggregations
    yoy = (
        master_df.groupby("Year")
        .agg(
            Total_Revenue   = ("Sales",  "sum"),
            Total_Profit    = ("Profit", "sum"),
            Total_Orders    = ("Order_ID", "nunique"),
        )
        .reset_index()
    )

    yoy["Profit_Margin_%"]      = (yoy["Total_Profit"] / yoy["Total_Revenue"] * 100).round(2)
    yoy["Avg_Order_Value"]      = (yoy["Total_Revenue"] / yoy["Total_Orders"]).round(2)

    # YoY growth % using shift()
    yoy["YoY_Revenue_Growth_%"] = (
        (yoy["Total_Revenue"] - yoy["Total_Revenue"].shift(1)) /
        yoy["Total_Revenue"].shift(1) * 100
    ).round(2)

    yoy["YoY_Profit_Growth_%"] = (
        (yoy["Total_Profit"] - yoy["Total_Profit"].shift(1)) /
        yoy["Total_Profit"].shift(1).abs() * 100
    ).round(2)

    # Best category per year
    best_cat = (
        master_df.groupby(["Year", "Category"])["Sales"]
        .sum()
        .reset_index()
        .sort_values(["Year", "Sales"], ascending=[True, False])
        .drop_duplicates("Year")
        .rename(columns={"Category": "Best_Category"})
    )
    yoy = yoy.merge(best_cat[["Year", "Best_Category"]], on="Year", how="left")

    # Best region per year
    best_reg = (
        master_df.groupby(["Year", "Region"])["Sales"]
        .sum()
        .reset_index()
        .sort_values(["Year", "Sales"], ascending=[True, False])
        .drop_duplicates("Year")
        .rename(columns={"Region": "Best_Region"})
    )
    yoy = yoy.merge(best_reg[["Year", "Best_Region"]], on="Year", how="left")

    # Save to Excel with formatting
    output_path = os.path.join(OUTPUTS_PATH, "yoy_trend_analysis.xlsx")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        yoy.to_excel(writer, index=False, sheet_name="YoY Trend")
        ws = writer.sheets["YoY Trend"]
        apply_excel_formatting(ws)

    print(f"  YoY analysis saved → {output_path}")
    print(yoy[["Year", "Total_Revenue", "YoY_Revenue_Growth_%"]].to_string(index=False))

    return yoy


# ============================================================
# STEP 5 — SUMMARY REPORT (MULTI-SHEET EXCEL)
# ============================================================

def step5_summary_report(master_df: pd.DataFrame, yoy_df: pd.DataFrame, quality_log: list):
    """Build a 6-sheet Excel summary report."""
    section_banner("STEP 5 — Building Summary Report")

    output_path = os.path.join(OUTPUTS_PATH, "summary_report.xlsx")

    # ── Sheet 2: By Category ────────────────────────────────
    by_cat = (
        master_df.groupby(["Year", "Category"])
        .agg(Revenue=("Sales", "sum"), Profit=("Profit", "sum"))
        .reset_index()
    )
    by_cat["Profit_Margin_%"] = (by_cat["Profit"] / by_cat["Revenue"] * 100).round(2)

    # ── Sheet 3: By Region ───────────────────────────────────
    by_region = (
        master_df.groupby(["Year", "Region"])
        .agg(Revenue=("Sales", "sum"), Orders=("Order_ID", "nunique"))
        .reset_index()
    )

    # ── Sheet 4: By Segment ──────────────────────────────────
    by_segment = (
        master_df.groupby(["Year", "Customer_Segment"])
        .agg(Revenue=("Sales", "sum"), Profit=("Profit", "sum"), Orders=("Order_ID", "nunique"))
        .reset_index()
    )

    # ── Sheet 5: Top 10 Products ─────────────────────────────
    top_products = (
        master_df.groupby("Product_Name")
        .agg(Total_Revenue=("Sales", "sum"), Total_Profit=("Profit", "sum"), Total_Orders=("Order_ID", "count"))
        .reset_index()
        .sort_values("Total_Revenue", ascending=False)
        .head(10)
    )
    top_products["Profit_Margin_%"] = (top_products["Total_Profit"] / top_products["Total_Revenue"] * 100).round(2)

    # ── Sheet 6: Data Quality Log ────────────────────────────
    quality_df = pd.DataFrame(quality_log)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Sheet 1
        yoy_df.to_excel(writer, index=False, sheet_name="YoY Trend")
        apply_excel_formatting(writer.sheets["YoY Trend"])

        # Sheet 2
        by_cat.to_excel(writer, index=False, sheet_name="By Category")
        apply_excel_formatting(writer.sheets["By Category"])

        # Sheet 3
        by_region.to_excel(writer, index=False, sheet_name="By Region")
        apply_excel_formatting(writer.sheets["By Region"])

        # Sheet 4
        by_segment.to_excel(writer, index=False, sheet_name="By Segment")
        apply_excel_formatting(writer.sheets["By Segment"])

        # Sheet 5
        top_products.to_excel(writer, index=False, sheet_name="Top Products")
        apply_excel_formatting(writer.sheets["Top Products"])

        # Sheet 6
        quality_df.to_excel(writer, index=False, sheet_name="Data Quality Log")
        apply_excel_formatting(writer.sheets["Data Quality Log"], header_fill_hex="C00000")

    print(f"  Summary report saved → {output_path}")
    print(f"  Sheets: YoY Trend | By Category | By Region | By Segment | Top Products | Data Quality Log")


# ============================================================
# STEP 6 — AUTOMATED CHARTS
# ============================================================

def step6_charts(master_df: pd.DataFrame, yoy_df: pd.DataFrame):
    """Generate and save 3 professional PNG charts."""
    section_banner("STEP 6 — Generating Charts")

    # ── Shared style ─────────────────────────────────────────
    sns.set_theme(style="whitegrid", font_scale=1.1)
    PALETTE = sns.color_palette("Blues_d", 12)

    # ── Chart 1: Revenue Trend (Line) ─────────────────────────
    fig, ax = plt.subplots(figsize=(12, 6))
    ax.plot(
        yoy_df["Year"], yoy_df["Total_Revenue"] / 1e6,
        marker="o", linewidth=2.5, color="#1F4E79", markersize=7
    )
    ax.fill_between(yoy_df["Year"], yoy_df["Total_Revenue"] / 1e6, alpha=0.12, color="#1F4E79")

    # Annotate each point
    for _, row in yoy_df.iterrows():
        ax.annotate(
            f"${row['Total_Revenue']/1e6:.1f}M",
            xy=(row["Year"], row["Total_Revenue"] / 1e6),
            xytext=(0, 10), textcoords="offset points",
            ha="center", fontsize=8.5, color="#1F4E79"
        )

    ax.set_title("Total Revenue Trend (2013–2024)", fontsize=16, fontweight="bold", pad=15)
    ax.set_xlabel("Year", fontsize=12)
    ax.set_ylabel("Revenue ($ Millions)", fontsize=12)
    ax.xaxis.set_major_locator(mticker.MultipleLocator(1))
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"${x:.1f}M"))
    ax.grid(axis="y", linestyle="--", alpha=0.5)
    ax.grid(axis="x", linestyle="", alpha=0)
    plt.tight_layout()
    path1 = os.path.join(CHARTS_PATH, "revenue_trend.png")
    plt.savefig(path1, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"  [1] revenue_trend.png saved → {path1}")

    # ── Chart 2: YoY Growth % (Bar) ───────────────────────────
    yoy_clean = yoy_df.dropna(subset=["YoY_Revenue_Growth_%"])

    colors = ["#C00000" if v < 0 else "#1F4E79" for v in yoy_clean["YoY_Revenue_Growth_%"]]

    fig, ax = plt.subplots(figsize=(12, 6))
    bars = ax.bar(yoy_clean["Year"], yoy_clean["YoY_Revenue_Growth_%"], color=colors, width=0.6)

    for bar, val in zip(bars, yoy_clean["YoY_Revenue_Growth_%"]):
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() + (0.3 if val >= 0 else -1.5),
            f"{val:.1f}%",
            ha="center", va="bottom", fontsize=9, fontweight="bold"
        )

    ax.axhline(0, color="black", linewidth=0.8, linestyle="--")
    ax.set_title("Year-over-Year Revenue Growth % (2014–2024)", fontsize=16, fontweight="bold", pad=15)
    ax.set_xlabel("Year", fontsize=12)
    ax.set_ylabel("Revenue Growth (%)", fontsize=12)
    ax.xaxis.set_major_locator(mticker.MultipleLocator(1))
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"{x:.0f}%"))
    ax.grid(axis="y", linestyle="--", alpha=0.5)
    plt.tight_layout()
    path2 = os.path.join(CHARTS_PATH, "yoy_growth.png")
    plt.savefig(path2, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"  [2] yoy_growth.png saved → {path2}")

    # ── Chart 3: Category Revenue Stacked Bar ─────────────────
    cat_pivot = (
        master_df.groupby(["Year", "Category"])["Sales"]
        .sum()
        .unstack(fill_value=0)
        / 1e6
    )

    cat_colors = {"Technology": "#1F4E79", "Furniture": "#2E75B6", "Office Supplies": "#9DC3E6"}
    fig, ax = plt.subplots(figsize=(13, 7))

    bottom = np.zeros(len(cat_pivot))
    for cat in cat_pivot.columns:
        color = cat_colors.get(cat, "#CCCCCC")
        bars = ax.bar(
            cat_pivot.index, cat_pivot[cat],
            bottom=bottom, label=cat,
            color=color, width=0.6
        )
        bottom += cat_pivot[cat].values

    ax.set_title("Revenue by Category per Year (2013–2024)", fontsize=16, fontweight="bold", pad=15)
    ax.set_xlabel("Year", fontsize=12)
    ax.set_ylabel("Revenue ($ Millions)", fontsize=12)
    ax.xaxis.set_major_locator(mticker.MultipleLocator(1))
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"${x:.1f}M"))
    ax.legend(title="Category", loc="upper left", framealpha=0.9)
    ax.grid(axis="y", linestyle="--", alpha=0.5)
    plt.tight_layout()
    path3 = os.path.join(CHARTS_PATH, "category_breakdown.png")
    plt.savefig(path3, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"  [3] category_breakdown.png saved → {path3}")


# ============================================================
# STEP 7 — LOAD INTO MYSQL
# ============================================================

def step7_load_mysql(master_df: pd.DataFrame):
    """
    Connect to MySQL, create the consolidated_sales table if needed,
    and bulk-insert the cleaned data.
    Uses mysql-connector-python (not SQLAlchemy).
    """
    section_banner("STEP 7 — Loading Data into MySQL")

    try:
        import mysql.connector
    except ImportError:
        print("  [SKIP] mysql-connector-python is not installed.")
        print("         Run: pip install mysql-connector-python")
        print("         Then re-run pipeline.py to complete this step.")
        return

    # ── Connect ──────────────────────────────────────────────
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor()
        print(f"  Connected to MySQL — database: {DB_CONFIG['database']}")
    except mysql.connector.Error as e:
        print(f"  [SKIP] Could not connect to MySQL: {e}")
        print("         Check your credentials in config.py and ensure the server is running.")
        return

    # ── Create database if not exists ────────────────────────
    try:
        cursor.execute(f"CREATE DATABASE IF NOT EXISTS `{DB_CONFIG['database']}`")
        cursor.execute(f"USE `{DB_CONFIG['database']}`")
    except mysql.connector.Error as e:
        print(f"  [ERROR] Could not create/use database: {e}")
        conn.close()
        return

    # ── Create table ─────────────────────────────────────────
    create_sql = """
    CREATE TABLE IF NOT EXISTS consolidated_sales (
        id                INT AUTO_INCREMENT PRIMARY KEY,
        Order_ID          VARCHAR(30),
        Order_Date        DATE,
        Customer_Name     VARCHAR(100),
        Customer_Segment  VARCHAR(50),
        Region            VARCHAR(50),
        Country           VARCHAR(50),
        Category          VARCHAR(50),
        Sub_Category      VARCHAR(50),
        Product_Name      VARCHAR(150),
        Quantity          INT,
        Unit_Price        DECIMAL(10,2),
        Discount          DECIMAL(5,2),
        Sales             DECIMAL(12,2),
        Profit            DECIMAL(12,2),
        Shipping_Cost     DECIMAL(10,2),
        Ship_Mode         VARCHAR(50),
        Year              INT,
        Profit_Margin_Pct DECIMAL(8,2),
        Revenue_Band      VARCHAR(10)
    )
    """
    cursor.execute(create_sql)
    print("  Table `consolidated_sales` ready.")

    # ── Truncate then insert (idempotent re-runs) ─────────────
    cursor.execute("TRUNCATE TABLE consolidated_sales")

    # Prepare DataFrame for insertion
    insert_df = master_df.copy()
    insert_df["Order_Date"] = insert_df["Order_Date"].dt.strftime("%Y-%m-%d")
    insert_df["Profit_Margin_Pct"] = insert_df["Profit_Margin_%"]
    insert_df["Revenue_Band"]       = insert_df["Revenue_Band"].astype(str)

    cols = [
        "Order_ID", "Order_Date", "Customer_Name", "Customer_Segment",
        "Region", "Country", "Category", "Sub_Category", "Product_Name",
        "Quantity", "Unit_Price", "Discount", "Sales", "Profit",
        "Shipping_Cost", "Ship_Mode", "Year", "Profit_Margin_Pct", "Revenue_Band"
    ]
    insert_df = insert_df[cols]

    insert_sql = f"""
        INSERT INTO consolidated_sales ({', '.join(cols)})
        VALUES ({', '.join(['%s'] * len(cols))})
    """

    data_tuples = [tuple(row) for row in insert_df.itertuples(index=False)]

    # Batch insert in chunks of 1000
    chunk_size = 1000
    total_inserted = 0
    for i in range(0, len(data_tuples), chunk_size):
        chunk = data_tuples[i : i + chunk_size]
        cursor.executemany(insert_sql, chunk)
        conn.commit()
        total_inserted += len(chunk)
        print(f"  Inserted rows: {total_inserted:,} / {len(data_tuples):,}", end="\r")

    print(f"\n  [OK] {total_inserted:,} rows inserted into `consolidated_sales`.")

    cursor.close()
    conn.close()
    print("  MySQL connection closed.")


# ============================================================
# MAIN — RUN PIPELINE END TO END
# ============================================================

def main():
    print("\n" + "█" * 60)
    print("  MULTI-YEAR SALES CONSOLIDATION PIPELINE")
    print("  2013 – 2024  |  python pipeline.py")
    print("█" * 60)

    try:
        # Step 1: Load
        yearly_dfs = step1_load_files(RAW_DATA_PATH)

        # Step 2: Clean
        cleaned_dfs, quality_log = step2_clean_all(yearly_dfs)

        # Step 3: Consolidate
        master_df = step3_consolidate(cleaned_dfs)

        # Step 4: YoY Analysis
        yoy_df = step4_yoy_analysis(master_df)

        # Step 5: Summary Report
        step5_summary_report(master_df, yoy_df, quality_log)

        # Step 6: Charts
        step6_charts(master_df, yoy_df)

        # Step 7: MySQL
        step7_load_mysql(master_df)

    except Exception as e:
        print(f"\n[FATAL ERROR] Pipeline failed:")
        traceback.print_exc()
        sys.exit(1)

    section_banner("PIPELINE COMPLETE")
    print("  Outputs generated:")
    print(f"    data/processed/consolidated_sales.csv")
    print(f"    outputs/yoy_trend_analysis.xlsx")
    print(f"    outputs/summary_report.xlsx")
    print(f"    outputs/charts/revenue_trend.png")
    print(f"    outputs/charts/yoy_growth.png")
    print(f"    outputs/charts/category_breakdown.png")
    print("\n  Manual consolidation time : 3 days")
    print("  Pipeline run time         : < 10 minutes")
    print("=" * 60)


if __name__ == "__main__":
    main()
